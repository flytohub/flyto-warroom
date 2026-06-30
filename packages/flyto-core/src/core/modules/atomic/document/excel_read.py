# Copyright 2026 Flyto2. Licensed under Apache-2.0. See LICENSE.

"""
Excel Read Module
Read data from Excel files (xlsx, xls)
"""
import logging
import os
from typing import Any, Dict, List, Optional

from ...registry import register_module
from ...schema import compose, presets


logger = logging.getLogger(__name__)


@register_module(
    module_id='excel.read',
    version='1.0.0',
    category='document',
    subcategory='excel',
    tags=['excel', 'spreadsheet', 'read', 'xlsx', 'data', 'path_restricted'],
    label='Read Excel',
    label_key='modules.excel.read.label',
    description='Read data from Excel files (xlsx, xls)',
    description_key='modules.excel.read.description',
    icon='Table',
    color='#217346',

    # Connection types
    input_types=['file_path'],
    output_types=['array', 'object'],
    can_connect_to=['array.*', 'data.*', 'database.*', 'flow.*'],
    can_receive_from=['file.*', 'data.*', 'http.*', 'flow.*', 'start'],

    # Execution settings
    timeout_ms=60000,
    retryable=False,
    concurrent_safe=True,

    # Security settings
    requires_credentials=False,
    handles_sensitive_data=True,
    required_permissions=['filesystem.read', 'filesystem.write'],

    params_schema=compose(
        presets.EXCEL_PATH(placeholder='/path/to/data.xlsx'),
        presets.EXCEL_SHEET(),
        presets.EXCEL_HEADER_ROW(),
        presets.EXCEL_RANGE(),
        presets.EXCEL_AS_DICT(),
    ),
    output_schema={
        'data': {
            'type': 'array',
            'description': 'Extracted data rows'
        ,
                'description_key': 'modules.excel.read.output.data.description'},
        'headers': {
            'type': 'array',
            'description': 'Column headers'
        ,
                'description_key': 'modules.excel.read.output.headers.description'},
        'row_count': {
            'type': 'number',
            'description': 'Number of data rows'
        ,
                'description_key': 'modules.excel.read.output.row_count.description'},
        'sheet_names': {
            'type': 'array',
            'description': 'All sheet names in the workbook'
        ,
                'description_key': 'modules.excel.read.output.sheet_names.description'}
    },
    examples=[
        {
            'title': 'Read entire sheet',
            'title_key': 'modules.excel.read.examples.basic.title',
            'params': {
                'path': '/tmp/data.xlsx',
                'as_dict': True
            }
        }
    ],
    author='Flyto Team',
    license='MIT'
)
async def excel_read(context: Dict[str, Any]) -> Dict[str, Any]:
    """Read data from Excel file"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel reading. Install with: pip install openpyxl")

    params = context['params']
    path = params['path']
    sheet_name = params.get('sheet')
    header_row = params.get('header_row', 1)
    cell_range = params.get('range')
    as_dict = params.get('as_dict', True)

    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    ws = _select_sheet(wb, sheet_name, sheet_names)

    all_rows = _read_rows(ws, cell_range)
    headers, data_rows = _extract_headers(all_rows, header_row)
    data = _rows_to_dicts(data_rows, headers) if as_dict and headers else data_rows

    wb.close()

    logger.info(f"Read Excel: {path} ({len(data)} rows)")

    return {
        'ok': True,
        'data': data,
        'headers': headers,
        'row_count': len(data),
        'sheet_names': sheet_names,
        'active_sheet': ws.title
    }


def _select_sheet(wb, sheet_name: Optional[str], sheet_names: List[str]):
    if sheet_name:
        if sheet_name not in sheet_names:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return wb[sheet_name]
    return wb.active


def _read_rows(ws, cell_range: Optional[str]) -> List[List[Any]]:
    cells = ws[cell_range] if cell_range else ws.iter_rows()
    return [[cell.value for cell in row] for row in cells]


def _extract_headers(all_rows: List[List[Any]], header_row: int):
    if header_row > 0 and len(all_rows) >= header_row:
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(all_rows[header_row - 1])]
        return headers, all_rows[header_row:]
    return [], all_rows


def _rows_to_dicts(data_rows: List[List[Any]], headers: List[str]) -> List[Dict[str, Any]]:
    result = []
    for row in data_rows:
        row_dict = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            row_dict[key] = val
        result.append(row_dict)
    return result
